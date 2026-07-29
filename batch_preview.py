# -*- coding: utf-8 -*-
"""단일 URL 또는 엑셀(링크 열) → lessons/previews/<slug>/<차시제목>.html."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from converter import convert_lesson
from offline_images import github_raw_url, lesson_slug, offline_lesson_data, preview_html_name
import app

ROOT = Path(__file__).resolve().parent
PREVIEWS = ROOT / "lessons" / "previews"
URL_RE = re.compile(r"https?://(?:www\.)?crayonschool\.co\.kr/lessons/[^\s\"'<>]+", re.I)


def normalize_url(url: str) -> str:
    u = (url or "").strip().rstrip(").,;]")
    if not u:
        return ""
    p = urlparse(u)
    if p.scheme not in ("http", "https"):
        return ""
    if "crayonschool.co.kr" not in (p.netloc or "").lower():
        return ""
    if "/lessons/" not in (p.path or ""):
        return ""
    return u


def urls_from_xlsx(path: Path) -> list[str]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("openpyxl 필요: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, data_only=True)
    found: list[str] = []
    seen: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                for m in URL_RE.findall(text):
                    u = normalize_url(m)
                    if u and u not in seen:
                        seen.add(u)
                        found.append(u)
                # 하이퍼링크만 있는 셀(표시 텍스트 ≠ URL)
                u = normalize_url(text)
                if u and u not in seen:
                    seen.add(u)
                    found.append(u)
    return found


def build_one(url: str) -> dict[str, Any]:
    url = normalize_url(url)
    if not url:
        return {"ok": False, "url": url, "error": "크레용스쿨 lessons URL이 아닙니다."}
    try:
        data = convert_lesson(url)
        title = data.get("title") or "lesson"
        slug = lesson_slug(title)
        html_name = preview_html_name(title)
        folder = PREVIEWS / slug
        folder.mkdir(parents=True, exist_ok=True)
        prefix = f"lessons/previews/{slug}"
        data = offline_lesson_data(data, folder, repo_prefix=prefix)
        html = app.build_html(data)
        out = folder / html_name
        out.write_text(html, encoding="utf-8")
        legacy = folder / "preview.html"
        if legacy.exists() and legacy.resolve() != out.resolve():
            try:
                legacy.unlink()
            except OSError:
                pass
        (folder / "meta.json").write_text(
            json.dumps(
                {
                    "title": data.get("title"),
                    "level": data.get("level"),
                    "sourceUrl": data.get("sourceUrl"),
                    "mode": data.get("mode"),
                    "pageCount": len(data.get("pages") or []),
                    "imagesOffline": True,
                    "htmlFile": html_name,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        raw = github_raw_url(f"{prefix}/{html_name}")
        blob = (
            "https://github.com/ABE0402/Lesson-Plan-Generator/blob/main/"
            f"{prefix}/{html_name}"
        )
        return {
            "ok": True,
            "url": url,
            "title": title,
            "slug": slug,
            "htmlFile": html_name,
            "path": out.as_posix(),
            "raw": raw,
            "blob": blob,
            "pageCount": len(data.get("pages") or []),
        }
    except Exception as e:  # noqa: BLE001 — 배치에서 한 건 실패해도 계속
        return {"ok": False, "url": url, "error": str(e)}


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="교안 preview 단건/엑셀 일괄 생성")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--url", help="크레용스쿨 차시 URL 1개")
    g.add_argument("--xlsx", type=Path, help="링크가 들어 있는 엑셀 파일")
    g.add_argument("--urls-file", type=Path, help="URL 한 줄씩 텍스트 파일")
    p.add_argument(
        "--limit",
        type=int,
        default=0,
        help="최대 N건만 (0=전부). 대량 요청 시 권장",
    )
    p.add_argument(
        "--course",
        default="",
        help="엑셀일 때 강좌명 부분 일치 필터(선택)",
    )
    args = p.parse_args(argv)

    urls: list[str] = []
    if args.url:
        urls = [args.url]
    elif args.xlsx:
        path = args.xlsx if args.xlsx.is_absolute() else (Path.cwd() / args.xlsx)
        if not path.exists():
            print(json.dumps({"ok": False, "error": f"파일 없음: {path}"}, ensure_ascii=False))
            return 1
        # 강좌 필터: openpyxl로 행 단위 재스캔
        if args.course:
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True)
            needle = args.course.strip().lower()
            seen: set[str] = set()
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if not any(needle in c.lower() for c in cells):
                        continue
                    for c in cells:
                        for m in URL_RE.findall(c):
                            u = normalize_url(m)
                            if u and u not in seen:
                                seen.add(u)
                                urls.append(u)
        else:
            urls = urls_from_xlsx(path)
    else:
        path = args.urls_file if args.urls_file.is_absolute() else (Path.cwd() / args.urls_file)
        for line in path.read_text(encoding="utf-8").splitlines():
            u = normalize_url(line.split("#", 1)[0].strip())
            if u:
                urls.append(u)

    if args.limit and args.limit > 0:
        urls = urls[: args.limit]

    if not urls:
        print(json.dumps({"ok": False, "error": "변환할 URL이 없습니다."}, ensure_ascii=False))
        return 1

    results = [build_one(u) for u in urls]
    ok = [r for r in results if r.get("ok")]
    fail = [r for r in results if not r.get("ok")]
    summary = {
        "ok": len(fail) == 0,
        "total": len(results),
        "success": len(ok),
        "failed": len(fail),
        "results": results,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
